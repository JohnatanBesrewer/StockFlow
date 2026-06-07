from openpyxl import load_workbook
import shutil

def print_compact_invoice_table(file_path, max_width=None):
    if max_width is None:
        max_width = min(120, shutil.get_terminal_size().columns - 2)

    wb = load_workbook(file_path)
    sheet = wb.active

    # Считываем все строки, заменяя None на пустую строку
    raw_rows = []
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            raw_rows.append([str(cell) if cell is not None else "" for cell in row])

    if not raw_rows:
        print("Нет данных для отображения.")
        return

    n_cols = max(len(row) for row in raw_rows)

    # Определяем, какие столбцы НЕ пустые
    non_empty_cols = []
    for col_idx in range(n_cols):
        if any(
            row[col_idx] != "" for row in raw_rows if col_idx < len(row)
        ):
            non_empty_cols.append(col_idx)

    if not non_empty_cols:
        print("Все столбцы пустые.")
        return

    # Формируем "очищенные" строки только с непустыми колонками
    clean_rows = []
    for row in raw_rows:
        clean_row = [row[i] if i < len(row) else "" for i in non_empty_cols]
        clean_rows.append(clean_row)

    # Определяем ширину колонок (минимум 1, максимум — чтобы влезло в экран)
    col_widths = []
    for i in range(len(non_empty_cols)):
        width = max(len(row[i]) for row in clean_rows)
        col_widths.append(max(1, width))

    # Уменьшаем ширину колонок, если общая ширина > max_width
    total_width = sum(col_widths) + 3 * (len(col_widths) - 1)  # " | " между колонками
    if total_width > max_width:
        # Простая стратегия: ограничить каждую колонку пропорционально
        # Но проще — ограничить максимум на колонку
        max_col_width = max(10, (max_width - 3 * (len(col_widths) - 1)) // len(col_widths))
        col_widths = [min(w, max_col_width) for w in col_widths]

    # Обрезаем текст, если он не влезает в колонку
    def truncate(text, width):
        if len(text) <= width:
            return text
        return text[:width - 1] + "…"

    # Форматирование строки
    def format_row(row):
        parts = []
        for i, cell in enumerate(row):
            w = col_widths[i]
            parts.append(truncate(cell, w).ljust(w))
        return " | ".join(parts)

    # Вывод
    for i, row in enumerate(clean_rows):
        if i == 1:  # после заголовка — разделитель
            sep = "-+-".join("-" * w for w in col_widths)
            print(sep)
        print(format_row(row))

# Пример использования
if __name__ == "__main__":
    print_compact_invoice_table("nakladnaya.xlsx")